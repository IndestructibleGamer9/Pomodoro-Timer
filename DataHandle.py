import matplotlib as plt
import pandas as pd 
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import Workbook
from datetime import datetime

USERNAME = 'TESTING'

class DataControl():
    def __init__(self):
        self.wb = None
        self.ws1 = None

    def connect(self, username='AnonymousUser'):
        filename = f'{username}.xlsx'
        try:
            self.wb = openpyxl.load_workbook(filename)
            self.ws1 = self.wb['Pomodoro Timer Data']
        except FileNotFoundError:
            print('File not found. Creating new file.')
            self.create_workbook(username)
        except InvalidFileException:
            print('Invalid file format. Creating new file.')
            self.create_workbook(username)
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            return

        # Find the last ID and create a new one
        last_id = 0
        for row in self.ws1.iter_rows(min_row=2, max_row=self.ws1.max_row, min_col=1, max_col=1):
            if row[0].value is not None and isinstance(row[0].value, int):
                last_id = max(last_id, row[0].value)

        new_id = last_id + 1

        # Find the first empty row
        self.new_row = self.ws1.max_row + 1
        while self.ws1.cell(row=self.new_row, column=1).value is not None:
            self.new_row += 1

        # Add new entry
        self.formatt()
        self.ws1.cell(row=self.new_row, column=1, value=new_id)
        self.ws1.cell(row=self.new_row, column=2, value=datetime.now())
        self.ws1.cell(row=self.new_row, column=5, value=True)  # Set as active

        # Save the changes
        self.wb.save(filename)

    def formatt(self):
        self.ws1 = self.wb['Pomodoro Timer Data']
        self.ws2= self.wb['Settings']
        self.ws3 = self.wb['Reminders']

        self.ws1['A1'] = 'ID'
        self.ws1['B1'] = 'Log On'
        self.ws1['C1'] = 'Log Off'
        self.ws1['D1'] = 'Promo Time'
        self.ws1['E1'] = 'Active'
        self.ws2['A1'] = 'sound'

    def create_workbook(self, name):
        self.wb = Workbook()
        fn = f'{name}.xlsx'

        # Format WB
        self.ws1 = self.wb.active
        self.ws1.title = "Pomodoro Timer Data"
        self.ws2 = self.wb.create_sheet(title='Settings')
        self.ws3 = self.wb.create_sheet(title='Reminders')

        self.ws1['A1'] = 'ID'
        self.ws1['B1'] = 'Log On'
        self.ws1['C1'] = 'Log Off'
        self.ws1['D1'] = 'Promo Time'
        self.ws1['E1'] = 'Active'
        self.ws2['A1'] = 'sound'

        self.wb.save(fn)

    def close(self, time, sound):
        fn = f'{USERNAME}.xlsx'
        self.ws1.cell(row=self.new_row, column=3, value=datetime.now())
        self.ws1.cell(row=self.new_row, column=4, value=time)
        self.ws2.cell(row=2, column=1, value=sound)
        self.wb.save(fn)


    def main(self):
        self.connect(USERNAME)

if __name__ == '__main__':
    data = DataControl()
    data.main()